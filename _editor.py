from openpyxl import Workbook, load_workbook, cell
from openpyxl.styles import *

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import tkcalendar
import tktreeviewtable

import os
import csv
import datetime

import sheetUtil
import datePicker

font_family = "TH Sarabun New"
normal_text = (font_family, '16')
easy_read = (font_family, '18')
input_font = (font_family, '14')
button_font = (font_family, '12', 'bold')
h1_font = (font_family, '32', 'bold')
h2_font = (font_family, '22', 'bold')
h3_font = (font_family, '20', 'bold')
hyperlink_font = (font_family, '14', 'bold underline')
small_para = (font_family, '12')


class Editor:

    def __init__(self, master, uniname="Student Attendant Manager - Testing Run", file="", class_num=None,
                 main_obj=None, new_obj=None):
        """ Constructor """

        self.no_file = False

        if file == "":
            file = "Workbook Example/Sample.xlsx"
            self.no_file = True

        if class_num is None:
            class_num = []

        self.UNINAME = uniname
        self.master = master
        self.workbook_name = os.path.basename(file)
        self.full_dir = file
        self.file = load_workbook(file)
        self.info_sheet = self.file["INFO"]
        self.current_sheet = None
        self.student_list = []
        self.sheet_info = {}
        self.date_dict = None
        self.class_num = class_num

        self.main_obj = main_obj
        self.new_obj = new_obj
        self.date_picker = None

        # ==========================================================================
        # ============================== START OF GUI ==============================
        # ==========================================================================

        self.main = tk.Toplevel(self.master)
        self.main.title("Editor | " + self.UNINAME)

        self.main.protocol("WM_DELETE_WINDOW", self.ask_save_quit)

        # ============================ STYLES =============================

        style = ttk.Style()
        style.configure("Treeview", font=("Tahoma", "10"))
        style.configure("Treeview.Heading", font=("Angsana New", "15", "bold"))
        style.configure("TLabel", font=normal_text)
        style.configure("Info.TLabel", relief=tk.SUNKEN, background="#FFF", font=input_font,
                        justify=tk.CENTER, anchor=tk.CENTER)

        # ============================ MENUs & CONFIGS =============================

        self.menu = tk.Menu(self.main)

        self.style = ttk.Style(self.main)
        self.style.configure("info.TLabel", background=[('pressed', '!disabled', 'black'), ('active', 'white')])

        self.menu.option_add("*Label.Font", normal_text)

        menu_file = tk.Menu(self.menu, tearoff=0)
        menu_edit = tk.Menu(self.menu, tearoff=0)
        menu_help = tk.Menu(self.menu, tearoff=0)

        menu_file.add_command(label="สร้างงานใหม่", command=self.new_file)
        menu_file.add_command(label="เปิด", command=self.open_file)
        menu_file.add_command(label="บันทึกงาน", command=self.save_override)
        menu_file.add_command(label="บันทึกเป็น...", command=self.save_as)
        menu_file.add_separator()
        menu_file.add_command(label="บันทึก และ ออก", command=self.save_and_quit)

        menu_edit.add_command(label="ลงข้อมูลนักเรียนหลายๆคน")
        menu_edit.add_separator()
        menu_edit.add_command(label="แก้ไขข้อมูลพื้นฐานของงาน")
        menu_edit.add_separator()
        menu_edit.add_command(label="แก้ไขกรอบเวลา")

        menu_help.add_command(label="เอกสารคู่มือ (Coming Soon)", state=tk.DISABLED)
        menu_help.add_command(label="เกี่ยวกับโปรแกรม")

        self.menu.add_cascade(label="ไฟล์", menu=menu_file)
        self.menu.add_cascade(label="การแก้ไข", menu=menu_edit)
        self.menu.add_cascade(label="ช่วยเหลือ", menu=menu_help)

        self.main.configure(menu=self.menu)

        self.main.bind("<Control-n>", lambda e: self.new_file())
        self.main.bind("<Control-o>", lambda e: self.open_file())
        self.main.bind("<Control-s>", lambda e: self.save_override())
        self.main.bind("<Control-Shift-s>", lambda e: self.save_as())

        # ============================= WORKING SCREEN =============================

        work_screen = ttk.Frame(self.main)
        work_screen.pack(side=tk.LEFT, padx=10, ipadx=25, pady=5)

        # --------------------------- Description Heading --------------------------

        desc_heading = ttk.LabelFrame(work_screen, text="ข้อมูลพื้นฐาน")
        desc_heading.pack()

        # **************** Left Part ****************
        desc_left = ttk.Frame(desc_heading)

        # Document Name
        doc_name_frame = ttk.Frame(desc_left)
        doc_name_frame.pack(pady=5, anchor=tk.W, expand=True, fill=tk.X)
        ttk.Label(doc_name_frame, text="ชื่อเอกสารรายงานนักเรียน: ").pack(side=tk.LEFT)
        self.book_name_lbl = ttk.Label(doc_name_frame, text=self.workbook_name, style="Info.TLabel", justify=tk.CENTER)
        self.book_name_lbl.pack(ipadx=25, expand=True, fill=tk.X)

        # Class & Affiliate
        class_frame = ttk.Frame(desc_left)
        class_frame.pack(pady=5, anchor=tk.W, expand=True, fill=tk.X)
        ttk.Label(class_frame, text="ระดับชั้น").pack(side=tk.LEFT)
        self.aff_lbl = ttk.Label(class_frame, text="AFF", style="Info.TLabel", justify=tk.CENTER)
        self.aff_lbl.pack(side=tk.LEFT, ipadx=10, expand=True, fill=tk.X)
        ttk.Label(class_frame, text="ปีที่ ").pack(side=tk.LEFT)
        self.class_lbl = ttk.Label(class_frame, text="G / R", style="Info.TLabel", justify=tk.CENTER)
        self.class_lbl.pack(ipadx=25, expand=True, fill=tk.X)

        # Subject & Hour
        subject_hour_frame = ttk.Frame(desc_left)
        subject_hour_frame.pack(pady=5, anchor=tk.W, expand=True, fill=tk.X)
        ttk.Label(subject_hour_frame, text="วิชา ").pack(side=tk.LEFT)
        self.subject_lbl = ttk.Label(subject_hour_frame, text="SUBJECT", style="Info.TLabel", justify=tk.CENTER)
        self.subject_lbl.pack(side=tk.LEFT, ipadx=10, padx=5, expand=True, fill=tk.X)
        ttk.Label(subject_hour_frame, text="จำนวนชั่วโมง: ").pack(side=tk.LEFT)
        self.total_hour_lbl = ttk.Label(subject_hour_frame, text="HOURS", style="Info.TLabel", justify=tk.CENTER)
        self.total_hour_lbl.pack(side=tk.LEFT, ipadx=5, padx=5, expand=True, fill=tk.X)

        # Date Frame
        dateframe_frame = ttk.Frame(desc_left)
        dateframe_frame.pack(pady=5, anchor=tk.W, expand=True, fill=tk.X)
        ttk.Label(dateframe_frame, text="ช่วงเวลาตั้งแต่ ").pack(side=tk.LEFT)
        self.date_from_lbl = ttk.Label(dateframe_frame, text="FROM", style="Info.TLabel", justify=tk.CENTER)
        self.date_from_lbl.pack(side=tk.LEFT, ipadx=5, expand=True, fill=tk.X)
        ttk.Label(dateframe_frame, text=" ถึง ").pack(side=tk.LEFT)
        self.date_to_lbl = ttk.Label(dateframe_frame, text="TO", style="Info.TLabel", justify=tk.CENTER)
        self.date_to_lbl.pack(side=tk.LEFT, ipadx=5, expand=True, fill=tk.X)

        # Teacher's Name
        teacher_name_frame = ttk.Frame(desc_left)
        ttk.Label(teacher_name_frame, text="ครูที่รับผิดชอบ:").pack(side=tk.LEFT)
        self.teacher_lbl = ttk.Label(teacher_name_frame, text="---- ----", style="Info.TLabel", justify=tk.CENTER)
        self.teacher_lbl.pack(side=tk.LEFT, ipadx=5, expand=True, fill=tk.X)
        teacher_name_frame.pack(pady=5, anchor=tk.W, expand=True, fill=tk.X)

        desc_left.pack(side=tk.LEFT)

        # **************** Right Part ****************
        desc_right = ttk.Frame(desc_heading)

        student_num_frame = ttk.Frame(desc_right)
        student_num_frame.pack()
        ttk.Label(student_num_frame, text="จำนวนนักเรียน", font=easy_read).pack(side=tk.LEFT)
        self.student_num_lbl = ttk.Label(student_num_frame, text="##", font=easy_read, justify=tk.CENTER,
                                         style="Info.TLabel")
        self.student_num_lbl.pack(side=tk.LEFT, padx=15, ipadx=5)
        ttk.Label(student_num_frame, text="คน", font=easy_read).pack(side=tk.LEFT)
        ttk.Button(desc_right, text="แก้ไขข้อมูลพื้นฐานของงาน").pack(expand=True, fill=tk.BOTH, pady=20)

        desc_right.pack(side=tk.LEFT, padx=10, pady=5, expand=True, fill=tk.Y)

        # ----------------------- Data Viewing Table ---------------------

        sheet_select = ttk.Frame(work_screen)
        ttk.Label(sheet_select, text="ชื่อแผ่นงาน").pack(side=tk.LEFT)
        self.sheet_list_select = ttk.Combobox(sheet_select)
        self.sheet_list_select.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=50)
        sheet_select.pack(expand=True, fill=tk.X, padx=25)

        self.viewtable = \
            tktreeviewtable.TreeviewTable(work_screen,
                                          header=["เลขที่", "รหัสประจำตัว", "ชื่อ-นามสกุล", "ขาด", "ลา", "ป่วย", "รวม"],
                                          enable_vsb=True, max_height=18, odd_row_color="#e5e5e5")
        self.viewtable.set_col_width([2, 5, 25, 0.01, 0.01, 0.01, 0.03])
        self.viewtable.container.pack(expand=True, fill=tk.X, padx=25, pady=5)

        # ================================ SEPARATOR ===============================

        ttk.Separator(self.main, orient=tk.VERTICAL).pack(side=tk.LEFT, expand=True, fill=tk.Y, pady=5)

        # Read the target workbook's main sheet and output information for editing.
        self.read_sheet()
        self.date_dict = sheetUtil.get_date_list(self.current_sheet)

        # ================================= TOOLBOX ================================

        toolbox = ttk.Frame(self.main)

        # ------------------------ Add Box -----------------------

        self.add_pane = AddPane(toolbox, self)
        self.add_pane.main.pack(expand=True, fill=tk.X, padx=20)

        # ----------------------- Search Box ---------------------

        self.search_pane = SearchPane(toolbox, self)
        self.search_pane.main.pack(expand=True, fill=tk.X, padx=20)

        # ----------------------- Edit Box -----------------------

        self.edit_pane = EditPane(toolbox, self, self.viewtable)
        self.edit_pane.set_pane_state("disabled")
        self.edit_pane.main.pack(expand=True, fill=tk.X, padx=20)

        toolbox.pack(side=tk.LEFT, expand=True, fill=tk.X)

        self.viewtable.bind_tree("<<TreeviewSelect>>", lambda e: self.edit_pane.show_selected())

        # ==========================================================================
        # =============================== END OF GUI ===============================
        # ==========================================================================

    def edit_sheet_info(self):
        # TODO: Open a description editor and pre-set the value for it
        #  (Copy the GUI from New File window, except for date-picking!)
        pass

    def read_sheet(self):
        """ Read the target workbook's main sheet and output information for editing. """
        name_reading = []
        for col in self.info_sheet.iter_cols(max_col=1, min_row=2):
            for row in col:
                if row.value is not None:
                    name_reading.append(row.value)
        self.sheet_info = {
            "name": name_reading.copy(),
            "aff": self.info_sheet["B2"].value,
            "grade": str(self.info_sheet["C2"].value),
            "room": str(self.info_sheet["D2"].value),
            "semester": self.info_sheet["E2"].value,
            "subject": self.info_sheet["F2"].value,
            "total_hrs": str(self.info_sheet["G2"].value),
            "teacher_name": self.info_sheet["H2"].value,
        }
        self.current_sheet = self.file[self.sheet_info["name"][0]]
        self.sheet_info["frame"] = sheetUtil.search_date_frame(self.current_sheet)
        self.student_list = sheetUtil.get_student_list(self.current_sheet, self.sheet_info["aff"])
        self.update_info_tab(self.sheet_info, self.sheet_info["name"][0])
        self.viewtable.set_content(self.student_list)

    def save_override(self):
        """ Save current worksheet with the same file name. (Typical saving) """
        try:
            self.file.save(self.full_dir)
        except PermissionError:
            print("PermissionError Flagged in save_override()")
            messagebox.showerror("ไม่สามารถบันทึกงานได้", "ไม่สามารถบันทึกงานได้\n"
                                                          "เนื่องจากท่านยังเปิดเอกสารแผ่นนี้ด้วยโปรแกรมอื่นอยู่\n"
                                                          "โปรดปิดโปรแกรมที่เปิดเอกสารนี้ก่อน "
                                                          "แล้วลองอีกครั้งภายหลัง")

    def save_as(self):
        """ Save current worksheet with different file name. (Save as) """
        new_filename = filedialog.asksaveasfilename(filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        try:
            load_workbook(new_filename)
        except FileNotFoundError:
            self.file.save(new_filename)
            self.change_file(new_filename)
        else:
            replace = messagebox.askyesno("พบไฟล์ซ้ำ", "ตรวจพบไฟล์ซ้ำ\n"
                                                       "ท่านต้องการบันทึกทับไฟล์ดังกล่าวหรือไม่")
            if replace:
                self.file.save(new_filename)
                self.change_file(new_filename)

    def save_and_quit(self):
        """ Save the current file, the return to the Main Menu. """
        self.save_override()
        self.main_obj.open()
        self.quit()

    def ask_save_quit(self):
        """ Ask for saving, save if yes; hen quit the editor """
        if messagebox.askyesno(title="บันทึกงาน", message="ท่านกำลังจะปิดงานนี้\n"
                                                          "บันทึกงานปัจจุบันไว้หรือไม่?"):
            self.save_override()
        self.quit()

    def change_file(self, file):
        """ Change the editor's current workbook. """
        self.workbook_name = os.path.basename(file)
        self.full_dir = file
        self.file = load_workbook(file)
        self.info_sheet = self.file["INFO"]
        self.current_sheet = None
        self.student_list = []
        self.sheet_info = {}

        self.book_name_lbl.config(text=self.workbook_name)
        self.read_sheet()

    def update_info_tab(self, info, target_sheet=""):
        """ Read the worksheet's description, and output those information onto the editor's Info Tab. """
        class_txt = info["grade"] + " / " + info["room"]
        self.aff_lbl.configure(text=info["aff"])
        self.class_lbl.configure(text=class_txt)
        self.subject_lbl.configure(text=info["subject"])
        self.total_hour_lbl.configure(text=info["total_hrs"])
        self.date_from_lbl.configure(text=info["frame"][0])
        self.date_to_lbl.configure(text=info["frame"][1])
        self.teacher_lbl.configure(text=info["teacher_name"])
        self.student_num_lbl.configure(text=str(len(self.student_list)))
        self.sheet_list_select.configure(value=info["name"])
        self.sheet_list_select.current(info["name"].index(target_sheet))

    def update_number_table(self):
        """ List down the list of students into the main table. """
        for num, student in enumerate(self.student_list, 1):
            student[0] = num
        self.viewtable.set_content(self.student_list)

    def update_student_list(self, name=None, absence=None):
        """ Get the list of some or all student's data. (With filtering) """
        self.student_list = sheetUtil.get_student_list(self.current_sheet, filter_name=name, filter_absence=absence)
        self.search_pane.search_num.configure(to=len(self.viewtable.tree.get_children()))
        self.viewtable.set_content(self.student_list)

    def link_to_main(self, widget):
        """ Self-explanatory; perform linking with Main Menu Widget for transmitting info. """
        self.main_obj = widget

    def link_to_new(self, widget):
        """ Self-explanatory; perform linking with New File Widget for transmitting info. """
        self.new_obj = widget

    def quit(self):
        self.main_obj.open()
        self.main.destroy()

    # def hide(self):
    #     self.main.withdraw()
    #
    # def show(self):
    #     self.main.update()
    #     self.main.deiconify()

    def open_file(self):
        if messagebox.askyesno(title="บันทึกงาน", message="ท่านกำลังจะปิดงานนี้\n"
                                                          "บันทึกงานปัจจุบันไว้หรือไม่?"):
            self.save_override()
        self.main_obj.openfile()
        self.quit()

    def new_file(self):
        if messagebox.askyesno(title="บันทึกงาน", message="ท่านกำลังจะปิดงานนี้\n"
                                                          "บันทึกงานปัจจุบันไว้หรือไม่?"):
            self.save_override()
        self.main_obj.open()
        self.main_obj.newfile()
        self.quit()


class SearchPane:

    def __init__(self, master, editor_obj):
        """ Constructor """
        self.master = master
        self.editor_obj = editor_obj

        self.main = ttk.Frame(self.master)

        # ==========================================================================
        # ============================== START OF GUI ==============================
        # ==========================================================================

        # -------------------------------------------------------------
        # ----------------------- Search by ID ------------------------
        # -------------------------------------------------------------

        search_by_id = ttk.LabelFrame(self.main, text="ค้นหาด้วยข้อมูลนักเรียน")

        # Number
        num_frame = ttk.Frame(search_by_id)
        ttk.Label(num_frame, text="เลขที่:").pack(side=tk.LEFT)
        self.search_num = ttk.Spinbox(num_frame, width=4, from_=0,
                                      to=len(self.editor_obj.viewtable.tree.get_children()))
        self.search_num.pack(side=tk.LEFT, expand=True, fill="x", padx=10)
        self.search_num.insert(0, "0")
        ttk.Button(num_frame, text="ค้นหา", command=lambda: self.search_id_callback(query="num")) \
            .pack(side=tk.RIGHT, anchor="e", padx=5, ipadx=5, pady=5)
        num_frame.pack(anchor=tk.W, expand=True, fill="x")

        # ID
        id_frame = ttk.Frame(search_by_id)
        ttk.Label(id_frame, text="เลขประจำตัว: ").pack(side=tk.LEFT)
        self.search_id = ttk.Spinbox(id_frame, from_=0, to=99999)
        self.search_id.pack(side=tk.LEFT, expand=True, fill="x", padx=10)
        self.search_id.insert(0, "0")
        ttk.Button(id_frame, text="ค้นหา", command=lambda: self.search_id_callback(query="id")) \
            .pack(side=tk.RIGHT, padx=5, ipadx=5, pady=5)
        id_frame.pack(anchor=tk.W, expand=True, fill="x")

        # Name
        name_frame = ttk.Frame(search_by_id)
        ttk.Label(name_frame, text="ชื่อ: ").pack(side=tk.LEFT)
        self.search_name = ttk.Entry(name_frame, width=30)
        self.search_name.pack(side=tk.LEFT, expand=True, fill="x", padx=10)
        # self.search_name.insert(0, "")
        ttk.Button(name_frame, text="ค้นหา", command=lambda: self.search_id_callback(query="name")) \
            .pack(side=tk.RIGHT, padx=5, ipadx=5, pady=5)
        name_frame.pack(anchor=tk.W, expand=True, fill="x")

        # Info Text
        ttk.Label(search_by_id, text="กรอกข้อมูลอย่างน้อย 1 ช่องเพื่อทำการค้นหา", relief=tk.GROOVE, font=small_para) \
            .pack(side="left", expand=True, fill=tk.X, padx=5)

        # Buttons
        ttk.Button(search_by_id, text="ล้างการค้นหา", command=self.reset_search) \
            .pack(side="left", expand=True, fill=tk.X, padx=5, pady=5)

        search_by_id.pack(side=tk.LEFT, ipadx=20, ipady=5, expand=True, fill=tk.BOTH)

        # -------------------------------------------------------------
        # -------------------- Search by frequency --------------------
        # -------------------------------------------------------------

        search_by_freq = ttk.LabelFrame(self.main, text="ค้นหาด้วยความถี่การไม่มาเรียน")
        ttk.Label(search_by_freq, text="ค้นตามจำนวนของ...").pack()

        absent_frame = ttk.Frame(search_by_freq)
        ttk.Label(absent_frame, text="ขาด: ").pack(side=tk.LEFT)
        self.search_absent = ttk.Spinbox(absent_frame, from_=0, to=100, width=3)
        self.search_absent.pack(side=tk.LEFT, padx=5)
        self.search_absent.insert(0, "0")
        ttk.Button(absent_frame, text="ค้นหา", command=lambda: self.search_freq(query="a")) \
            .pack(side=tk.LEFT, padx=5)
        absent_frame.pack(anchor=tk.E)

        on_duty_frame = ttk.Frame(search_by_freq)
        ttk.Label(on_duty_frame, text="ลา: ").pack(side=tk.LEFT)
        self.search_on_duty = ttk.Spinbox(on_duty_frame, from_=0, to=100, width=3)
        self.search_on_duty.pack(side=tk.LEFT, padx=5)
        self.search_on_duty.insert(0, "0")
        ttk.Button(on_duty_frame, text="ค้นหา", command=lambda: self.search_freq(query="o")) \
            .pack(side=tk.LEFT, padx=5)
        on_duty_frame.pack(anchor=tk.E)

        ill_frame = ttk.Frame(search_by_freq)
        ttk.Label(ill_frame, text="ป่วย: ").pack(side=tk.LEFT)
        self.search_ill = ttk.Spinbox(ill_frame, from_=0, to=100, width=3)
        self.search_ill.pack(side=tk.LEFT, padx=5)
        self.search_ill.insert(0, "0")
        ttk.Button(ill_frame, text="ค้นหา", command=lambda: self.search_freq(query="i")) \
            .pack(side=tk.LEFT, padx=5)
        ill_frame.pack(anchor=tk.E)

        not_present_frame = ttk.Frame(search_by_freq)
        ttk.Label(not_present_frame, text="ทั้งหมด: ").pack(side=tk.LEFT)
        self.search_not_present = ttk.Spinbox(not_present_frame, from_=0, to=100, width=3)
        self.search_not_present.pack(side=tk.LEFT, padx=5)
        self.search_not_present.insert(0, "0")
        ttk.Button(not_present_frame, text="ค้นหา", command=lambda: self.search_freq(query="n")) \
            .pack(side=tk.LEFT, padx=5)
        not_present_frame.pack(anchor=tk.E)

        search_by_freq.pack(side=tk.LEFT, fill=tk.Y, ipadx=5, ipady=5)

        # ==========================================================================
        # =============================== END OF GUI ===============================
        # ==========================================================================

    def search_id_callback(self, query="num"):
        try:
            query_num, query_id, query_name = int(self.search_num.get()), int(self.search_id.get()), \
                                              self.search_name.get()
        except ValueError:
            messagebox.showerror("เกิดข้อผิดพลาด", "โปรดกรอกข้อมูลในการค้นหาด้วยค่ะ")
            return
        else:
            for child in self.editor_obj.viewtable.tree.get_children():
                if query == "num" and \
                        int(self.editor_obj.viewtable.tree.item(child)["values"][0]) == query_num:
                    self.editor_obj.viewtable.tree.selection_set(child)
                    return
                elif query == "id" and \
                        int(self.editor_obj.viewtable.tree.item(child)["values"][1]) == query_id:
                    self.editor_obj.viewtable.tree.selection_set(child)
                    return
            if query == "name":
                self.editor_obj.update_student_list(name=query_name)
            else:
                messagebox.showerror("เกิดข้อผิดพลาด", "ไม่พบนักเรียนที่มีข้อมูลดังกล่าว โปรดใช้ตัวกรองการค้นหาอื่นๆ")
                return

    def search_freq(self, query="a"):
        try:
            switch = {"a": int(self.search_absent.get()), "o": int(self.search_on_duty.get()),
                      "i": int(self.search_ill.get()), "n": int(self.search_not_present.get())}
        except ValueError:
            messagebox.showerror("เกิดข้อผิดพลาด", "โปรดกรอกข้อมูลในการค้นหาด้วยค่ะ")
            return
        self.editor_obj.update_student_list(absence=(query, switch[query]))

    def reset_search(self):
        self.search_num.set(0)
        self.search_id.set(0)
        self.search_name.config(text="")
        self.editor_obj.update_student_list()


class EditPane:

    def __init__(self, master, editor_obj, viewtable):
        self.editor_obj = editor_obj
        self.viewtable = viewtable
        self.master = master
        self.main = ttk.LabelFrame(self.master, text="แก้ไขรายคน")
        self.date_dict = sheetUtil.get_date_list(self.editor_obj.current_sheet)
        self.month_list = self.date_dict.keys()
        self.student_index = 0
        self.absence_index = 0

        # ==========================================================================
        # ============================== START OF GUI ==============================
        # ==========================================================================

        # ---------- Info Bar -----------

        info_box = ttk.Frame(self.main)
        info_box.pack(expand=True, fill=tk.X, padx=25)

        # ----- Info Tab -----
        student_info = ttk.Frame(info_box)
        self.lbl1 = ttk.Label(student_info, text="ชื่อ:")
        self.lbl1.pack(side=tk.LEFT)
        self.name = ttk.Label(student_info, text="N/A", style="Info.TLabel", justify=tk.CENTER, font=input_font)
        self.name.pack(side=tk.LEFT, expand=True, fill=tk.X, ipadx=20)
        self.lbl2 = ttk.Label(student_info, text="  เลขประจำตัว:")
        self.lbl2.pack(side=tk.LEFT)
        self.id = ttk.Label(student_info, text="N/A", style="Info.TLabel", justify=tk.CENTER, font=input_font)
        self.id.pack(side=tk.LEFT, expand=True, fill=tk.X)
        self.lbl3 = ttk.Label(student_info, text="  เลขที่:")
        self.lbl3.pack(side=tk.LEFT)
        self.num = ttk.Label(student_info, text="N/A", style="Info.TLabel", justify=tk.CENTER, font=input_font)
        self.num.pack(side=tk.LEFT, ipadx=5)
        student_info.pack(anchor=tk.W, expand=True, fill=tk.X, padx=10)

        # ----- Absence Tab -----
        absence_info = ttk.Frame(info_box)

        self.lbl4 = ttk.Label(absence_info, text="ขาด:")
        self.lbl4.pack(side=tk.LEFT)
        self.a_num = ttk.Label(absence_info, text="N/A", style="Info.TLabel", justify=tk.CENTER, font=input_font)
        self.a_num.pack(side=tk.LEFT, ipadx=10)

        self.lbl5 = ttk.Label(absence_info, text="  ลา:")
        self.lbl5.pack(side=tk.LEFT)
        self.o_num = ttk.Label(absence_info, text="N/A", style="Info.TLabel", justify=tk.CENTER, font=input_font)
        self.o_num.pack(side=tk.LEFT, ipadx=10)

        self.lbl6 = ttk.Label(absence_info, text="  ป่วย:")
        self.lbl6.pack(side=tk.LEFT)
        self.i_num = ttk.Label(absence_info, text="N/A", style="Info.TLabel", justify=tk.CENTER, font=input_font)
        self.i_num.pack(side=tk.LEFT, ipadx=10)

        absence_info.pack(anchor=tk.W, expand=True, fill=tk.X, pady=5, padx=10)

        self.edit_std_info = ttk.Button(absence_info, text="แก้ไขข้อมูลนักเรียน")
        self.edit_std_info.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=20)

        # Absence Table
        self.absence_table = tktreeviewtable \
            .TreeviewTable(self.main, header=["วันที่ไม่มาเรียน", "ประเภท"], max_height=5, enable_vsb=True,
                           odd_row_color="#e5e5e5")
        self.absence_table.set_col_width([12, 3])
        self.absence_table.container.pack(expand=True, fill=tk.X, padx=25, pady=5)

        # Editing Form
        edition_frame = ttk.Frame(self.main)
        form = ttk.Frame(edition_frame)
        self.form_lbl1 = ttk.Label(form, text="วันที่:")
        self.form_lbl2 = ttk.Label(form, text="  เดือน:")
        self.form_lbl3 = ttk.Label(form, text="  ประเภท:")
        self.date_entry = ttk.Combobox(form, value=("-",), width=2)
        self.month_entry = ttk.Combobox(form, value=("N/A",), width=10)
        self.type_entry = ttk.Combobox(form, value=("ขาด", "ลา", "ป่วย"), width=4)

        self.form_lbl1.pack(side=tk.LEFT)
        self.date_entry.pack(side=tk.LEFT)
        self.form_lbl2.pack(side=tk.LEFT)
        self.month_entry.pack(side=tk.LEFT)
        self.form_lbl3.pack(side=tk.LEFT)
        self.type_entry.pack(side=tk.LEFT)
        form.pack()

        # Button Sections
        self.add_only = ttk.Frame(edition_frame)
        self.add_btn = ttk.Button(self.add_only, text="เพี่ม", command=self.add_absence)
        self.add_btn.pack(expand=True, fill=tk.X)
        self.edit_remove = ttk.Frame(edition_frame)
        self.edit_btn = ttk.Button(self.edit_remove, text="แก้ไข", command=self.edit_absence)
        self.remove_btn = ttk.Button(self.edit_remove, text="นำออก", command=self.remove_absence)
        self.cancel_btn = ttk.Button(self.edit_remove, text="ยกเลิก", command=lambda: self.set_editing_mode("add"))
        self.edit_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=10)
        self.remove_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=10)
        self.cancel_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=10)

        self.set_editing_mode()
        edition_frame.pack(expand=True, fill=tk.X, padx=25)

        self.absence_table.bind_tree("<<TreeviewSelect>>", lambda e: self.absence_selection_callback())

        # ==========================================================================
        # =============================== END OF GUI ===============================
        # ==========================================================================

    def absence_selection_callback(self):
        try:
            self.absence_table.selection_callback()
            self.set_editing_mode("edit")
            date = int(self.absence_table.current_selection["values"][0].split(" / ")[0])
            month = self.absence_table.current_selection["values"][0].split(" / ")[1]
            abs_type = self.absence_table.current_selection["values"][1]

            date_index = self.date_dict[month].index(date)
            month_index = list(self.month_list).index(month)
            if abs_type in "ขาด":
                type_index = 0
            elif abs_type in "ลา":
                type_index = 1
            else:
                type_index = 2

            self.month_entry.current(month_index)
            self.set_date_selection_callback(self.month_entry, self.date_entry)
            self.date_entry.current(date_index)
            self.type_entry.current(type_index)
        except IndexError:
            messagebox.showerror("เกิดข้อผิดพลาด", "ไม่สามารถเลือกวันที่นี้ได้ โปรดลองอีกครั้งภายหลัง")
            return

    def set_pane_state(self, state):
        """ Set the state of the whole widget. """
        self.lbl1.configure(state=state)
        self.lbl2.configure(state=state)
        self.lbl3.configure(state=state)
        self.lbl4.configure(state=state)
        self.lbl5.configure(state=state)
        self.lbl6.configure(state=state)
        self.name.configure(state=state)
        self.id.configure(state=state)
        self.num.configure(state=state)
        self.a_num.configure(state=state)
        self.o_num.configure(state=state)
        self.i_num.configure(state=state)
        self.edit_std_info.configure(state=state)
        self.absence_table.set_widget_state(state)
        self.form_lbl1.configure(state=state)
        self.form_lbl2.configure(state=state)
        self.form_lbl3.configure(state=state)
        self.date_entry.configure(state=state)
        self.month_entry.configure(state=state)
        self.type_entry.configure(state=state)
        self.add_btn.configure(state=state)
        self.edit_btn.configure(state=state)
        self.remove_btn.configure(state=state)
        self.cancel_btn.configure(state=state)

    def set_editing_mode(self, state="add"):
        """ Set the editing button's mode """
        if state in "add":
            self.add_only.pack(expand=True, fill=tk.X, pady=5)
            self.edit_remove.pack_forget()
        elif state in "edit":
            self.add_only.pack_forget()
            self.edit_remove.pack(expand=True, fill=tk.X, pady=5)

    def show_selected(self):
        self.viewtable.selection_callback()
        self.set_pane_state("normal")
        selected = self.viewtable.current_selection["values"].copy()
        self.name.config(text=selected[2])
        self.id.config(text=selected[1])
        self.num.config(text=selected[0])
        self.student_index = selected[0] - 1
        if self.editor_obj.sheet_info["aff"] in "ประถมศึกษา":
            start_point = self.editor_obj.current_sheet["B7"]
        else:
            start_point = self.editor_obj.current_sheet["B8"]
        match_number = sheetUtil.search_top_down(self.editor_obj.current_sheet, start_point, selected[0])
        absence_list = sheetUtil.get_absence(self.editor_obj.current_sheet, match_number)
        a, o, i = 0, 0, 0
        for absence in absence_list:
            if absence[1] in "ขาด":
                a += 1
            elif absence[1] in "ลา":
                o += 1
            elif absence[1] in "ป่วย":
                i += 1
        self.a_num.config(text=str(a))
        self.o_num.config(text=str(o))
        self.i_num.config(text=str(i))
        self.absence_table.set_content(absence_list)

        month_choice = []
        for date_set in self.date_dict.items():
            month_choice.append(date_set[0])
        self.month_entry.configure(values=month_choice)
        self.month_entry.bind("<<ComboboxSelected>>",
                              lambda e: self.set_date_selection_callback(self.month_entry, self.date_entry))
        self.month_entry.current(0)
        self.type_entry.current(0)
        self.set_date_selection_callback(self.month_entry, self.date_entry)

    def set_date_selection_callback(self, month_combo, date_combo):
        date_combo.configure(value=self.date_dict[month_combo.get()])
        date_combo.current(0)

    def add_absence(self):
        for item_id in self.absence_table.tree.get_children():
            if str(self.date_entry.get()) + " / " + self.month_entry.get() in \
                    self.absence_table.tree.item(item_id, option="value"):
                messagebox.showinfo("มีข้อมูลซ้ำแล้ว", "มีข้อมูลการหยุดเรียนที่ตรงกับวันที่นี้\n"
                                                       "โปรดแก้ไขข้อมูลของวันที่นี้แทน")
                self.absence_table.tree.selection_set(item_id)
                self.set_editing_mode("edit")
                return
        if self.editor_obj.sheet_info["aff"] in "ประถมศึกษา":
            down_start_cell = self.editor_obj.current_sheet["B7"]
        else:
            down_start_cell = self.editor_obj.current_sheet["B8"]
        if self.type_entry.get() in "ขาด":
            type_letter = "ข"
        elif self.type_entry.get() in "ลา":
            type_letter = "ล"
        else:
            type_letter = "ป"
        row_cell = sheetUtil.search_top_down(self.editor_obj.current_sheet, down_start_cell, data=self.num.cget("text"))
        col_cell = sheetUtil.search_in_narrow_group(self.editor_obj.current_sheet,
                                                    group=self.month_entry.get(),
                                                    data=self.date_entry.get(),
                                                    start="E5")
        target = sheetUtil.intersect_cells(self.editor_obj.current_sheet, row_cell, col_cell)
        target.value = type_letter
        self.editor_obj.update_student_list()
        self.editor_obj.viewtable.tree.selection_set(self.editor_obj.viewtable.tree.get_children()[self.student_index])

    def remove_absence(self):
        self.absence_table.selection_callback()
        selected = self.absence_table.current_selection["values"].copy()
        selected_month = selected[0].split(" / ")[1]
        selected_date = selected[0].split(" / ")[0]
        if self.editor_obj.sheet_info["aff"] in "ประถมศึกษา":
            down_start_cell = self.editor_obj.current_sheet["B7"]
        else:
            down_start_cell = self.editor_obj.current_sheet["B8"]
        row_cell = sheetUtil.search_top_down(self.editor_obj.current_sheet, down_start_cell, data=self.num.cget("text"))
        col_cell = sheetUtil.search_in_narrow_group(self.editor_obj.current_sheet,
                                                    group=selected_month,
                                                    data=selected_date,
                                                    start="E5")
        target = sheetUtil.intersect_cells(self.editor_obj.current_sheet, row_cell, col_cell)
        target.value = None
        self.set_editing_mode("add")
        self.editor_obj.update_student_list()
        self.absence_table.tree.delete(self.absence_table.tree.selection())

    def edit_absence(self):
        self.remove_absence()
        self.add_absence()


class AddPane:

    def __init__(self, master, editor_obj):

        self.master = master
        self.editor_obj = editor_obj

        # ==========================================================================
        # ============================== START OF GUI ==============================
        # ==========================================================================

        self.main = ttk.LabelFrame(self.master, text="เพี่มนักเรียน")

        form = ttk.Frame(self.main)
        ttk.Label(form, text="เลขประจำตัว:").pack(side=tk.LEFT)
        self.id_entry = ttk.Spinbox(form, width=5, from_=1, to=99999)
        self.id_entry.pack(side=tk.LEFT)
        ttk.Label(form, text="  ชื่อ-นามสกุล:").pack(side=tk.LEFT)
        self.name_entry = ttk.Entry(form)
        self.name_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(form, text="เพี่ม", command=self.add_single).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=25)
        form.pack(expand=True, fill=tk.X)

        import_frame = ttk.Frame(self.main)
        ttk.Button(import_frame, text="นำเข้ารายชื่อจากไฟล์ Excel / CSV", command=self.import_names).pack(side=tk.LEFT)
        link = ttk.Label(import_frame, text="เรียนรู้เพี่มเติม", font=hyperlink_font, foreground="blue", cursor="hand2")
        link.pack(side=tk.LEFT)
        import_frame.pack()

        # link.bind("<Button-1>", """Open How-to Top-level widget""")

        # ==========================================================================
        # =============================== END OF GUI ===============================
        # ==========================================================================

    def add_single(self, import_info=None):

        # Check if the program address import information
        if import_info is None:
            ref_id = int(self.id_entry.get())
            ref_name = self.name_entry.get()
        else:
            ref_id = int(import_info[0])
            ref_name = import_info[1]

        if ref_id is None or ref_name is None:
            print("Not enough information.")
            return
        else:
            try:
                int(ref_id)
            except ValueError:
                print("Invalid input has been entered.")
                return
            else:
                result_num = 0
                for num, student in enumerate(self.editor_obj.student_list, 1):
                    if int(student[1]) > ref_id:
                        result_num = num
                        break
                    elif int(student[1]) == ref_id:
                        print("Duplicate ID has been entered. Consider editing the current student info.")
                        return
                if result_num == 0:
                    result_num = 1
                self.editor_obj.student_list.append([result_num, str(ref_id), ref_name, 0, 0, 0, 0])
                self.editor_obj.student_list.sort(key=lambda e: int(e[1]))
                self.editor_obj.update_number_table()
                sheetUtil.add_one_student(self.editor_obj.current_sheet,
                                          ref_id, ref_name)

    def import_names(self):

        file = tk.filedialog.askopenfilename(initialdir="Document", title="Select file",
                                             filetypes=[("xlsx files", "*.xlsx"),
                                                        ("csv files", "*.csv")])
        if file is not None:
            extension = os.path.splitext(file)
            if extension[1] in ".xlsx":
                book = load_workbook(file)
                mainsheet = book.active
                for row in mainsheet.iter_rows(max_col=2):
                    self.add_single(import_info=[row[0].value, row[1].value])
            elif extension[1] in ".csv":
                data_list = csv.reader(open(file, "r"))
                for data in data_list:
                    self.add_single(import_info=[data[0], data[1]])


if __name__ == '__main__':
    root = tk.Tk()
    editor = Editor(root)
    root.main.mainloop()
