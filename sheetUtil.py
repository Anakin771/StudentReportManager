import tkinter as tk
import tkinter.messagebox

import datetime

from openpyxl import Workbook, load_workbook, cell
from colorama import Fore, Style
from openpyxl.styles import *

"""
This sheet is used for Excel working specific to this project.
Any function(s) or constant(s) that revolve working into Excel files
should be contained in this document.

Author - Yanakorn Chaeyprasert
"""

month_to_thai = {
    "January": "มกราคม",
    "February": "กุมพาพันธ์",
    "March": "มีนาคม",
    "April": "เมษายน",
    "May": "พฤษภาคม",
    "June": "มิถุนายน",
    "July": "กรกฎาคม",
    "August": "สิงหาคม",
    "September": "กันยายน",
    "October": "ตุลาคม",
    "November": "พฤศจิกายน",
    "December": "ธันวาคม"
}

month_to_eng = {
    "มกราคม": "January",
    "กุมพาพันธ์": "February",
    "มีนาคม": "March",
    "เมษายน": "April",
    "พฤษภาคม": "May",
    "มิถุนายน": "June",
    "กรกฎาคม": "July",
    "สิงหาคม": "August",
    "กันยายน": "September",
    "ตุลาคม": "October",
    "พฤศจิกายน": "November",
    "ธันวาคม": "December"
}

weekday_to_thai = {
    "Mon": "จันทร์", "Tue": "อังคาร", "Wed": "พุธ", "Thu": "พฤหัสบดี", "Fri": "ศุกร์"
}

weekday_sort = {
    "Mon": 1, "Tue": 2, "Wed": 3, "Thu": 4, "Fri": 5
}

black_side = Side(border_style="thin", color="000000")
normal_edge = Border(left=black_side,
                     right=black_side,
                     bottom=black_side,
                     top=black_side)


def merge_len(worksheet, start):
    """Find a length of a merged cells"""
    start_cell = worksheet[start]
    merge_length = 1
    finished = False
    for rw in worksheet.iter_rows(min_row=start_cell.row, max_row=start_cell.row, min_col=start_cell.column):
        for clmn in rw:
            if isinstance(clmn, cell.Cell):
                main_cell = clmn
                while not finished:
                    offset_cell = main_cell.offset(0, merge_length)
                    if isinstance(offset_cell, cell.MergedCell):
                        merge_length += 1
                    else:
                        finished = True
            if finished:
                break
        if finished:
            break
    return merge_length


def detect_type_in_row(worksheet, start, report=False):
    """
    Allows us to search for position and type of each cell.
    (MergedCell do not count, however, its origin cell's position is located as a representative)
    """
    start_cell = worksheet[start]
    merge_index = []
    normal_index = []
    for rw in worksheet.iter_rows(min_col=start_cell.column, min_row=start_cell.row, max_row=start_cell.row):
        for clmn in rw:
            if isinstance(clmn, cell.Cell):
                lower_cell = clmn.offset(1, 0)
                right_cell = clmn.offset(0, 1)
                if isinstance(lower_cell, cell.MergedCell) or isinstance(right_cell, cell.MergedCell):
                    merge_index.append(clmn.coordinate)
                else:
                    normal_index.append(clmn.coordinate)
    if report:
        result = detect_type_in_row(worksheet, start)
        string_normal_index = "N/A"
        if len(result[0]) > 0:
            string_normal_index = ""
            for indx in result[0]:
                string_normal_index += indx + ", "
            string_normal_index = string_normal_index[:-2]
        string_merge_index = "N/A"
        if len(result[1]) > 0:
            string_merge_index = ""
            for indx in result[1]:
                string_merge_index += indx + ", "
            string_merge_index = string_merge_index[:-2]

        print("Founded " + str(len(result[0])) + " Non-Merge cell(s), at index: " + string_normal_index)
        print("Founded " + str(len(result[1])) + " Merge cell(s), at index: " + string_merge_index)
    return [normal_index, merge_index]


def search_in_narrow_group(worksheet, group, data, start, report=False):
    """
    Use this function whenever you want to search for a cell,
    which may contain the same value as the other on the same row,
    but is distinguished by a Merged Cell on top of it.

    :param worksheet: The worksheet file
    :param group: The value of the merged cell above the target cell
    :param data: The value of the target cell.
    :param start: The index of the beginning of search
    :param report: Optional; if set to true, a console will print the result before returning
    :return: Cell object which matches all criteria given above, or None if there are no matches.
    """
    start_cell = worksheet[start]
    found_group = False
    group_match = -1
    for rw in worksheet.iter_rows(min_row=start_cell.row, max_row=start_cell.row, min_col=start_cell.column):
        for clmn in rw:
            if clmn.value == group and isinstance(clmn, cell.Cell):
                group_match = clmn
                found_group = True
            if found_group:
                break
        if found_group:
            break
    if group_match == -1:
        print(Fore.RED + "Cannot find group '" + str(group) + "' (Error thrown from search_in_narrow_group())"
              + Style.RESET_ALL)
        return None
    group_limit = merge_len(worksheet, group_match.coordinate)
    lower_cell = group_match.offset(1, 0)
    search_cell = 0
    search_count = 1
    found_cell = False
    for rw in worksheet.iter_rows(min_row=lower_cell.row, max_row=lower_cell.row, min_col=lower_cell.column):
        for clmn in rw:
            if search_count <= group_limit and not found_cell:
                if str(clmn.value) == str(data):
                    search_cell = clmn
                    found_cell = True
                else:
                    search_count += 1
            elif not found_cell:
                print(Fore.RED + "Cannot find the cell (not group) with the data '" + str(data)
                      + "' (Error thrown from search_in_narrow_group())" + Style.RESET_ALL)
                return None
    if report:
        print(Fore.CYAN + "Found target cell at position " + search_cell.coordinate + Style.RESET_ALL)
    return search_cell


def search_top_down(sheet, start_cell, data):
    """
    Find the cell in ONE column (starting from start_cell's position downward)
    that has the value matches the data argument.

    :param sheet: Target worksheet to be searched.
    :param start_cell: Mandatory; specifies the starting cell. (input the argument as a Cell object)
    :param data: Data which you want to collate.
    :return: Cell object which has its value match with the data argument, or None, if no match is found.
    """
    result_cell = None
    for col in sheet.iter_cols(min_col=start_cell.column, max_col=start_cell.column, min_row=start_cell.row):
        for row in col:
            if row.value == data:
                result_cell = row
                return result_cell
    if result_cell is None:
        print(Fore.RED + "Error: Cannot find matching cell."
              + Style.RESET_ALL)
        return None


def intersect_cells(sheet, row_cell, col_cell):
    """
    Find an intersect of two cells in the same worksheet.

    :param sheet: Target worksheet to be searched.
    :param row_cell: A cell in the worksheet 'sheet' that specify the cell's row.
    :param col_cell: A cell in the worksheet 'sheet' that specify the cell's column.
    :return: A cell within the worksheet 'sheet' which has the same row in 'row_cell' and same column as 'col_cell'.
    """
    row = str(row_cell.row)
    col = col_cell.column_letter
    return sheet[col + row]


def get_absence(sheet, row_cell):
    """
    (Prior to show_selected function in _editor.py)
    Obtain the list of dates and types of absence.

    :param sheet: Target worksheet to be read.
    :param row_cell: Target cell which its row indexed the target row that will be read.
    :return: A 2D List which specifies the date of absence and type of absence; as follows:
             [
                [Date-0, Type-0],
                [Date-1, Type-1],
                        .
                        .
                        .
                [Date-n, Type-n]
             ]
    """
    result_list = []
    for row in sheet.iter_rows(min_col=5, max_col=26, min_row=row_cell.row, max_row=row_cell.row):
        for col in row:
            if col.value == "ข":
                date_absence = find_date_of_absence(sheet, col)
                result_list.append([date_absence, "ขาด"])
            elif col.value == "ล":
                date_absence = find_date_of_absence(sheet, col)
                result_list.append([date_absence, "ลา"])
            elif col.value == "ป":
                date_absence = find_date_of_absence(sheet, col)
                result_list.append([date_absence, "ป่วย"])
    return result_list


def get_date_list(sheet, aff):
    """
    (Prior to show_selected function in _editor.py)
    find the dates and months list of the given sheet.

    :param sheet: Target worksheet to be read.
    :param aff: The affiliate of the worksheet.
    :return: A dictionary; consists of multiple month-keys and dates-values list as follow:
            {
                "Month-0": [D0, D1, D2, ... Dn]-0,
                "Month-1": [D0, D1, D2, ... Dn]-1,
                                .
                                .
                                .
                "Month-n": [D0, D1, D2, ... Dn]-n,
            }
    """
    result = {}

    # Step 1: Find Month and assign them into list:
    if aff in "ประถมศึกษา":
        months = sheet["E6":"Z6"]
    else:
        months = sheet["E5":"Z5"]

    months_list = []
    for row in months:
        for col in row:
            if isinstance(col, cell.Cell):
                if col.value is not None:
                    months_list.append(col.value)

    # Step 2: Find Dates and assign them into 2D list:
    dates = sheet["E6":"Z6"]
    dates_list = []
    dates_sublist = []
    for row in dates:
        for col in row:
            if col.value is not None and col.offset(column=1).value is not None:
                if int(col.offset(column=1).value) <= int(col.value):
                    dates_sublist.append(col.value)
                    dates_list.append(dates_sublist.copy())
                    dates_sublist.clear()
                else:
                    dates_sublist.append(col.value)
            elif col.coordinate == "Z6":
                dates_sublist.append(col.value)
    dates_list.append(dates_sublist.copy())
    dates_sublist.clear()

    # Step 3: Merge each information into a single dictionary:
    for indx, month in enumerate(months_list):
        result[month] = dates_list[indx]
    return result


def find_date_of_absence(sheet, index_cell):
    """
    (Prior to get_absence and get_class_num function in this document)
    Obtain the date of absence.

    :param sheet: Target worksheet to be read.
    :param index_cell: Target cell; which will lead to the date header.
    :return: A string; specify the date of absence which corresponds to that of the target cell.
    """
    date = sheet[index_cell.column_letter + "6"].value
    month_search = sheet[index_cell.column_letter + "6"]
    look_left = 0
    while not isinstance(month_search.offset(row=-1, column=look_left), cell.Cell):
        look_left -= 1
    month_result = month_search.offset(row=-1, column=look_left).value
    return str(date) + " / " + month_result


def format_date_frame(dateframe_list, sheet, aff="ประถมศึกษา"):
    """
    (Prior to generate_sheet function in this document and datePicker.py uses)
    Print date data into target worksheet.

    :param dateframe_list: The compiled data of the date frame.
    :param sheet: Worksheet you want to read.
    :param aff: The affiliate of the class in that sheet.
    :return: Nothing; but it results in printed date in the worksheet.
    """

    if aff in "ประถมศึกษา":
        next_merge, next_date_fill = "E6", "E7"
        next_hrs_fill = None
    else:
        next_merge, next_date_fill, next_hrs_fill = "E5", "E6", "E7"

    for frame in dateframe_list:
        start_of_merge = sheet[next_merge]
        end_of_merge = start_of_merge.offset(column=len(frame["dates"]) - 1)
        next_merge = end_of_merge.offset(column=1).coordinate
        sheet.merge_cells(start_of_merge.coordinate + ':' + end_of_merge.coordinate)
        start_of_merge.value = month_to_thai[frame["month"]]
        start_of_merge.border = normal_edge
        for date in frame["dates"]:
            date_print = sheet[next_date_fill]
            date_print.value = int(date)
            date_print.alignment = Alignment(horizontal='center', vertical='center')
            next_date_fill = date_print.offset(column=1).coordinate
        if aff in "มัฐยมศึกษา":
            for class_order in frame["class_num"]:
                hrs_print = sheet[next_hrs_fill]
                hrs_print.value = int(class_order)
                next_hrs_fill = hrs_print.offset(column=1).coordinate


def generate_sheet(name="New Sheet", location='', dated_frame=None, grd=1, rm=1, aff="ประถมศึกษา", term=1,
                   smstr=str(int(datetime.datetime.now().strftime("%Y")) + 543),
                   hour=100, sbj="คณิตศาสตร์", s_unit=1, t_name="อาจารย์ประจำวิชา"):
    """
    (Prior to NewFile Class's use in _newSheet.py)
    Generate workbook based on given data (or with default data if wishes)

    :param name: The name of the workbook.
    :param location: Directory you wishes to be saved in.
    :param dated_frame: A Dateframe list, which contains dictionaries of what dates & months
                        to be printed on the worksheet.
    :param grd: Workbook's classroom grade.
    :param rm: Workbook's classroom room number.
    :param aff: Workbook's class affiliate (meaning, Elementary / Mid-School / High-School etc.)
    :param term: Term of the semester.
    :param smstr: Workbook's semester of work.
    :param hour: Total hours in the course.
    :param sbj: Workbook's subject of the class
    :param s_unit: Score Unit of that subject.
    :param t_name: Name of the responsible teacher.
    :return: Nothing; except it creates a new Excel Workbook with each information printed.
    """

    # Default Value Handling
    if dated_frame is None:
        dated_frame = []

    # Load Template Workbook
    if aff in "มัฐยมศึกษา":
        book = load_workbook("_TEMPLATE_M.xlsx")
        a3 = "ของนักเรียนชั้น" + aff + "ปีที่  " + str(grd) + " / " + str(rm) + " ภาคเรียนที่ " + str(term) +\
             "  ปีการศึกษา  " + smstr
        b4 = "วิชา \t " + sbj + " \t\t จำนวน \t " + str(s_unit) + "\t หน่วยกิต"
    else:
        book = load_workbook("_TEMPLATE_P.xlsx")
        a3 = "ของนักเรียนชั้น" + aff + "ปีที่  " + str(grd) + " / " + str(rm) + "  ปีการศึกษา  " + smstr
        b4 = "วิชา \t " + sbj + " \t\t จำนวน \t " + str(hour) + "\t ชั่วโมง"
    sheet = book["MAIN"]
    sheet.title = str(grd) + "_" + str(rm) + "_1"

    # Pin down data into info sheet
    info = book["TEMP_INFO"]
    info.title = "INFO"
    info["A2"].value = str(grd) + "_" + str(rm) + "_1"
    info["B2"].value = aff
    info["C2"].value = grd
    info["D2"].value = rm
    info["E2"].value = term
    info["F2"].value = smstr
    info["G2"].value = sbj
    info["H2"].value = s_unit
    info["I2"].value = hour
    info["J2"].value = t_name

    # Assigning Data
    sheet["A3"].value = a3
    sheet["B4"].value = b4
    sheet["Q4"].value = \
        "จำนวนนักเรียน \t 0 \t คน"
    if aff in "ประถมศึกษา":
        sheet["E5"].value = "ภาคเรียนที่ " + str(term)
        sheet["L11"].value = "(" + t_name + ")"
    else:
        sheet["L11"].value = "(" + t_name + ")"

    # Assign data into TEMP sheet
    temp = book["TEMP"]

    temp["A3"].value = a3
    temp["B4"].value = b4
    temp["Q4"].value = \
        "จำนวนนักเรียน \t 0 \t คน"
    if aff in "ประถมศึกษา":
        temp["E5"].value = "ภาคเรียนที่ " + str(term)
        temp["L11"].value = "(" + t_name + ")"
    else:
        temp["L11"].value = "(" + t_name + ")"

    # Applying Date Frame into MAIN sheet
    format_date_frame(dated_frame, sheet, aff)

    # Finalizing
    if location == '':
        book.save(name + ".xlsx")
    else:
        location = location + "/"
        book.save(location + name + ".xlsx")
    # Testing if sheet is successfully generated
    try:
        load_workbook(location + name + ".xlsx")
    except FileNotFoundError:
        tk.messagebox.showerror("เกิดข้อผิดพลาด", "เกิดข้อผิดพลาดในการสร้างไฟล์งานขึ้นมาใหม่\n"
                                                  "โปรดลองอีกครั้งภายหลังหรือติดต่อผู้ที่เกี่ยวข้อง")
    else:
        tk.messagebox.showinfo("การสร้างไฟล์สำเร็จ", "ไฟล์ของท่านสร้างสำเร็จแล้ว\n"
                                                     "กำลังนำท่านไปยังหน้า Editor")


def search_date_frame(sheet):
    """
    (Prior to Editor.read_sheet() in _editor.py)
    Find the first and last date of the sheet.

    :param sheet: Target worksheet to be searched.
    :return: A tuple of the first and the last date listed on that sheet.
    """
    first_date = str(sheet["E6"].value) + " " + str(sheet["E5"].value)
    current_last_month = ""
    for rw in sheet.iter_rows(min_row=5, max_row=5, min_col=5):
        for clmn in rw:
            if clmn.value is not None and isinstance(clmn, cell.Cell):
                current_last_month = clmn.value
    last_date = str(sheet["Z6"].value) + " " + str(current_last_month)
    return first_date, last_date


def get_student_list(sheet, filter_name=None, filter_absence=None):
    """
    (Prior to Editor class in _editor.py)
    Read and compile students' information on the target worksheet.

    :param sheet: Target worksheet to which student's to be read.
    :param filter_name: Optional, specify the target name for searching query.
    :param filter_absence: Optional, specify the type of absence and frequency of that absence.
    :return: The compiled result list of some or all student in the worksheet as follows:
             [
                [Number-0, ID-0, Name-0, Absence-0, Off-to-Duty-0, Ill-0, All-0],
                [Number-1, ID-1, Name-1, Absence-1, Off-to-Duty-1, Ill-1, All-1],
                                            .
                                            .
                                            .
                [Number-n, ID-n, Name-n, Absence-n, Off-to-Duty-n, Ill-n, All-n]
             ]
    """
    start_cell = sheet["B8"]
    result = []
    for col in sheet.iter_cols(min_col=start_cell.column, max_col=start_cell.column, min_row=start_cell.row):
        for row in col:
            if row.value is None:
                return result
            current_id = []
            a, o, i, n = 0, 0, 0, 0
            for row_id in sheet.iter_rows(min_col=row.column, max_col=4, min_row=row.row, max_row=row.row):
                for col_id in row_id:
                    current_id.append(col_id.value)
            for row_abs in sheet.iter_rows(min_col=5, max_col=26, min_row=row.row, max_row=row.row):
                for col_abs in row_abs:
                    if str(col_abs.value) in "ข" and not str(col_abs.value) == "":
                        a += 1
                        n += 1
                    elif str(col_abs.value) in "ล" and not str(col_abs.value) == "":
                        o += 1
                        n += 1
                    elif str(col_abs.value) in "ป" and not str(col_abs.value) == "":
                        i += 1
                        n += 1

            switch_abs = {"a": a, "o": o, "i": i, "n": n}
            if filter_name is None and filter_absence is None:
                result.append([current_id[0], current_id[1], current_id[2], a, o, i, n])
            elif filter_name is not None and filter_name in current_id[2]:
                # Search for specified text in the name.
                result.append([current_id[0], current_id[1], current_id[2], a, o, i, n])
            elif filter_absence is not None and filter_absence[1] == int(switch_abs[filter_absence[0]]):
                # Search for specified frequency of absence.
                result.append([current_id[0], current_id[1], current_id[2], a, o, i, n])


def add_one_student(workbook, id_num, name):
    """
    (Prior to AddStudent class in _editor.py)
    Add a row of one student into the sheet

    :param workbook: Target document to be added.
    :param id_num: Student's ID number to be added.
    :param name:Student's Name to be added.
    :return: Nothing; but it inserts extra row at the right index with the given information printed on it.
    """
    # TODO: Add new student to all student worksheet including template.

    for sheet_row in workbook["INFO"].iter_rows(min_row=2, min_col=1, max_col=1):
        for sheet_col in sheet_row:

            sheet = workbook[sheet_col.value]

            def print_info():
                current_coord = row.coordinate
                sheet.insert_rows(row.row)
                id_cell = sheet[current_coord]
                name_cell = id_cell.offset(column=1)
                id_cell.value = int(id_num)
                name_cell.value = name
                return sheet[current_coord]

            start_cell = sheet["C8"]
            for col in sheet.iter_cols(min_col=start_cell.column, max_col=start_cell.column, min_row=start_cell.row):
                for row in col:
                    if row.value is None:
                        current = print_info()
                        update_student_num(sheet)
                        update_styling(sheet, current)
                        return
                    elif int(id_num) < int(row.value):
                        current = print_info()
                        update_student_num(sheet)
                        update_styling(sheet, current)
                        return


def update_student_num(sheet):
    """
    (Prior to add_one_student in this document)
    Rerun the sheet number to correspond the order in the sheet's table.

    :param sheet: Target worksheet to be rerun.
    :return: Nothing; but the target worksheet will have its student's number redeemed correctly.
    """

    start_cell = sheet["B8"]
    count = 1
    for col in sheet.iter_cols(min_col=start_cell.column, max_col=start_cell.column, min_row=start_cell.row):
        for row in col:
            if row.offset(column=1).value is None:
                return
            else:
                row.value = count
                count += 1


def update_styling(sheet, target_cell):
    """
    (Prior to add_one_student in this document)
    Style the newly added row with correct styling fashion.

    :param sheet: Target worksheet to be styled.
    :param target_cell: Any cell that has its row indexed to the newly added row.
    :return: Nothing; but that new row of the worksheet is decorated according to its affiliate.
    """
    for row in sheet.iter_rows(min_col=2, max_col=26, min_row=target_cell.row, max_row=target_cell.row):
        for col in row:
            if col.column == 2:
                col.font = Font(name="Angsana New", size=14)
                col.alignment = Alignment(horizontal="center", vertical="center")
            elif col.column == 3:
                col.font = Font(name="Angsana New", size=16)
                col.alignment = Alignment(horizontal="center", vertical="center")
            else:
                col.font = Font(name="Angsana New", size=14)
            col.border = normal_edge


def extend_sheet(workbook, aff, s_class, dateframe):
    s_num = 1
    for col in workbook["INFO"].iter_cols(min_row=2, min_col=1, max_col=1):
        for row in col:
            if row.value is not None:
                s_num += 1
            else:
                row.value = "{}_{}".format(s_class, str(s_num))
    ext = workbook.copy_worksheet(workbook["TEMP"])
    ext.title = "{}_{}".format(s_class, str(s_num))
    format_date_frame(dateframe, ext, aff)


def close_doc(workbook, full):
    # TODO: "Close" the document via cut the last column of the last sheet and display each student's present hours
    #  on that column.
    info = workbook["INFO"]
    for col in info.iter_cols(min_row=2, min_col=1, max_col=1):
        for row in col:
            if row.offset(row=1).value is None:
                row.value = "{}_END".format(row.value)
